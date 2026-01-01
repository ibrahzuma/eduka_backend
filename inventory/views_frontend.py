from django.views.generic import ListView, CreateView, TemplateView, View, UpdateView, DeleteView
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image as PDFImage, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT
from reportlab.lib.units import inch, cm
from datetime import datetime
from django.conf import settings
import os
from django.shortcuts import redirect, render
from django.http import HttpResponse
import csv
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.contrib import messages
from .models import Product, Category, Stock, StockMovement
from .utils import generate_pdf_labels
from .forms import ProductForm, CategoryForm, StockAdjustmentForm, StockTransferForm, PurchaseForm
from shops.models import Shop, Branch
import io
from django.db import transaction
from django.db.models import F, Sum, Count, Case, When, Value, DecimalField
from django.core.exceptions import ValidationError
from decimal import Decimal

class BaseShopView(LoginRequiredMixin):
    """Base view to handle shop context and security"""
    
    def get_shop(self):
        # 1. Direct check for Employee's assigned shop
        if getattr(self.request.user, 'shop', None):
            return self.request.user.shop
            
        # 2. Check for Owner's shops
        if hasattr(self.request.user, 'shops') and self.request.user.shops.exists():
            return self.request.user.shops.first()
            
        # 3. Legacy fallback
        if hasattr(self.request.user, 'employee_profile'):
             return self.request.user.employee_profile.shop
             
        return None

class ProductListView(BaseShopView, ListView):
    model = Product
    template_name = 'inventory/product_list.html'
    context_object_name = 'products'

    def get_queryset(self):
        shop = self.get_shop()
        if shop:
            return Product.objects.filter(shop=shop)
        return Product.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'product_form' not in context:
            context['product_form'] = ProductForm(shop=self.get_shop())
        return context

    def post(self, request, *args, **kwargs):
        form = ProductForm(request.POST, request.FILES, shop=self.get_shop())
        if form.is_valid():
            shop = self.get_shop()
            if not shop:
                 messages.error(request, "No shop associated with account.")
                 return redirect('product_list')
            
            product = form.save(commit=False)
            product.shop = shop
            product.save()

            # Handle Stock Initialization (Opening Stock)
            # define opening stock for the CURRENT USER'S (or default) branch?
            # For now, let's update ALL branches or just the first/primary one.
            # Ideally, we should know which branch the user is operating in.
            # Since create_initial_stock signal creates 0 stock, we just update it.
            
            opening_stock = form.cleaned_data.get('opening_stock', 0)
            threshold = form.cleaned_data.get('low_stock_threshold', 0)
            
            # Update stocks for this product
            stocks = Stock.objects.filter(product=product)
            if stocks.exists():
                stocks.update(quantity=opening_stock, low_stock_threshold=threshold)
            
            messages.success(request, "Product created successfully!")
            return redirect('product_list')
        
        # If form is invalid
        self.object_list = self.get_queryset()
        context = self.get_context_data(product_form=form)
        return self.render_to_response(context)

class ProductCreateView(BaseShopView, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('product_list')

    def get_form_kwargs(self):
        kwargs = super(ProductCreateView, self).get_form_kwargs()
        kwargs['shop'] = self.get_shop()
        return kwargs

    def form_valid(self, form):
        shop = self.get_shop()
        if not shop:
             messages.error(self.request, "No shop associated with account.")
             return self.form_invalid(form)
        form.instance.shop = shop
        messages.success(self.request, "Product created successfully!")
        return super().form_valid(form)

class ProductUpdateView(BaseShopView, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('product_list')

    def get_form_kwargs(self):
        kwargs = super(ProductUpdateView, self).get_form_kwargs()
        kwargs['shop'] = self.get_shop()
        return kwargs

    def get_queryset(self):
        # Security: Only allow editing products in own shop
        shop = self.get_shop()
        if shop:
            return Product.objects.filter(shop=shop)
        return Product.objects.none()

    def form_valid(self, form):
        messages.success(self.request, "Product updated successfully!")
        return super().form_valid(form)

class ProductDeleteView(BaseShopView, DeleteView):
    model = Product
    template_name = 'inventory/product_confirm_delete.html'
    success_url = reverse_lazy('product_list')

    def get_queryset(self):
         # Security: Only allow deleting products in own shop
        shop = self.get_shop()
        if shop:
            return Product.objects.filter(shop=shop)
        return Product.objects.none()

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, "Product deleted successfully!")
        return super().delete(request, *args, **kwargs)

class CategoryListView(BaseShopView, ListView):
    model = Category
    template_name = 'inventory/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        shop = self.get_shop()
        if shop:
            return Category.objects.filter(shop=shop)
        return Category.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = CategoryForm()
        return context

    def post(self, request, *args, **kwargs):
        form = CategoryForm(request.POST)
        if form.is_valid():
            shop = self.get_shop()
            if not shop:
                 messages.error(request, "No shop associated with account.")
                 return redirect('category_list')
            
            form.instance.shop = shop
            form.save()
            messages.success(request, "Category created successfully!")
            return redirect('category_list')
        
        # If form is invalid, re-render the list with the form containing errors
        self.object_list = self.get_queryset()
        context = self.get_context_data(form=form)
        return self.render_to_response(context)

# Placeholders for others for now
# Removed Placeholders as they are now fully implemented below

class ServiceListView(BaseShopView, ListView):
    model = Product
    template_name = 'inventory/product_list.html'
    context_object_name = 'products'

    def get_queryset(self):
        shop = self.get_shop()
        if shop:
            return Product.objects.filter(shop=shop, product_type=Product.Type.SERVICE)
        return Product.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Services'
        if 'product_form' not in context:
            # Initialize form with Service type
            context['product_form'] = ProductForm(
                shop=self.get_shop(), 
                initial={'product_type': Product.Type.SERVICE}
            )
        return context

    def post(self, request, *args, **kwargs):
        form = ProductForm(request.POST, request.FILES, shop=self.get_shop())
        if form.is_valid():
            shop = self.get_shop()
            if not shop:
                 messages.error(request, "No shop associated with account.")
                 return redirect('service_list')
            
            product = form.save(commit=False)
            product.shop = shop
            # Force type to Service regardless of form input if on service page
            product.product_type = Product.Type.SERVICE 
            product.save()
            
            messages.success(request, "Service created successfully!")
            return redirect('service_list')
        
        # If form is invalid
        self.object_list = self.get_queryset()
        context = self.get_context_data(product_form=form)
        return self.render_to_response(context)

class ServiceCreateView(BaseShopView, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_form.html'
    success_url = reverse_lazy('service_list')

    def get_form_kwargs(self):
        kwargs = super(ServiceCreateView, self).get_form_kwargs()
        kwargs['shop'] = self.get_shop()
        return kwargs

    def get_initial(self):
        return {'product_type': Product.Type.SERVICE}
    
    def form_valid(self, form):
        shop = self.get_shop()
        if not shop:
             messages.error(self.request, "No shop associated with account.")
             return self.form_invalid(form)
        form.instance.shop = shop
        messages.success(self.request, "Service created successfully!")
        return super().form_valid(form)

class StockListView(BaseShopView, ListView):
    model = Stock
    template_name = 'inventory/stock_list.html'
    context_object_name = 'stocks'

    def get_queryset(self):
        shop = self.get_shop()
        if shop:
            return Stock.objects.filter(branch__shop=shop).select_related('product', 'branch')
        return Stock.objects.none()

    def post(self, request, *args, **kwargs):
        form = StockAdjustmentForm(request.POST)
        if form.is_valid():
            stock_id = form.cleaned_data['stock_id']
            adj_type = form.cleaned_data['adjustment_type']
            qty = form.cleaned_data['quantity']
            
            try:
                stock = Stock.objects.get(id=stock_id, branch__shop=self.get_shop())
                
                if adj_type == 'ADD':
                    stock.quantity += qty
                    change = qty
                    m_type = StockMovement.Type.ADD
                elif adj_type == 'REDUCE':
                    change = -qty
                    m_type = StockMovement.Type.REDUCE
                    stock.quantity = max(0, stock.quantity - qty)
                elif adj_type == 'SET':
                    change = qty - stock.quantity
                    m_type = StockMovement.Type.SET
                    stock.quantity = qty
                
                stock.save()
                
                # Log Movement
                StockMovement.objects.create(
                    stock=stock,
                    product=stock.product,
                    branch=stock.branch,
                    quantity_change=change,
                    movement_type=m_type,
                    reason=form.cleaned_data.get('reason'),
                    user=request.user
                )

                messages.success(request, f"Stock updated successfully for {stock.product.name}.")
                
            except Stock.DoesNotExist:
                messages.error(request, "Stock record not found or access denied.")
            except Exception as e:
                messages.error(request, f"Error updating stock: {str(e)}")
                
            return redirect('stock_list')
        
        # If form invalid
        messages.error(request, "Invalid form data.")
        return redirect('stock_list')

class StockManagementView(BaseShopView, ListView):
    model = StockMovement
    template_name = 'inventory/stock_movement_list.html'
    context_object_name = 'movements'
    paginate_by = 20

    def get_queryset(self):
        shop = self.get_shop()
        if shop:
            return StockMovement.objects.filter(branch__shop=shop).select_related('product', 'branch', 'user').order_by('-created_at')
def export_stock_pdf(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    shop = None
    if getattr(request.user, 'shop', None):
        shop = request.user.shop
    elif hasattr(request.user, 'shops') and request.user.shops.exists():
        shop = request.user.shops.first()
    elif hasattr(request.user, 'employee_profile'):
        shop = request.user.employee_profile.shop
    
    if not shop:
        return redirect('stock_list')

    stocks = Stock.objects.filter(branch__shop=shop).select_related('product', 'branch')

    response = HttpResponse(content_type='application/pdf')
    filename = f"Stock_Report_{datetime.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Document Setup
    doc = SimpleDocTemplate(
        response, 
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    elements = []
    styles = getSampleStyleSheet()

    # Custom Styles
    styles.add(ParagraphStyle(name='TitleStyle', parent=styles['Heading1'], alignment=TA_CENTER, fontSize=18, textColor=colors.HexColor('#4e73df'), spaceAfter=12))
    styles.add(ParagraphStyle(name='SubtitleStyle', parent=styles['Normal'], alignment=TA_CENTER, fontSize=12, textColor=colors.grey, spaceAfter=24))
    styles.add(ParagraphStyle(name='TableHeader', parent=styles['Normal'], fontSize=10, textColor=colors.white, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='TableData', parent=styles['Normal'], fontSize=9, textColor=colors.black))
    styles.add(ParagraphStyle(name='TableNumber', parent=styles['Normal'], fontSize=9, textColor=colors.black, alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER))

    # --- Header Section ---
    # Logo
    logo_path = os.path.join(settings.BASE_DIR, 'eduka_backend', 'static', 'img', 'logoeduka.png')
    if os.path.exists(logo_path):
        try:
            # Create a table for the header: Logo Left, Text Right (or Center)
            # Actually, a centered layout often looks cleaner for reports
            im = PDFImage(logo_path, width=2.0*inch, height=0.8*inch)
            im.hAlign = 'CENTER'
            elements.append(im)
            elements.append(Spacer(1, 12))
        except Exception as e:
            print(f"PDF Logo Error: {e}")

    # Title & Info
    elements.append(Paragraph(f"Stock Inventory Report", styles['TitleStyle']))
    elements.append(Paragraph(f"<b>Shop:</b> {shop.name} | <b>Date:</b> {datetime.now().strftime('%B %d, %Y')}", styles['SubtitleStyle']))

    # --- Table Section ---
    # Headers
    headers = [
        Paragraph('Product Name', styles['TableHeader']),
        Paragraph('Branch', styles['TableHeader']),
        Paragraph('Qty', styles['TableHeader']),
        Paragraph('Cost Price', styles['TableHeader']),
        Paragraph('Total Value', styles['TableHeader']),
    ]
    
    data = [headers]
    total_val = 0
    total_items = 0

    # Data Rows
    for stock in stocks:
        val = stock.quantity * (stock.product.cost_price or 0)
        total_val += val
        total_items += stock.quantity
        
        row = [
            Paragraph(stock.product.name[:35], styles['TableData']), # Truncate check
            Paragraph(stock.branch.name[:20], styles['TableData']),
            Paragraph(str(stock.quantity), styles['TableNumber']),
            Paragraph(f"{stock.product.cost_price:,.2f}", styles['TableNumber']),
            Paragraph(f"{val:,.2f}", styles['TableNumber']),
        ]
        data.append(row)

    # Footer Row (Totals)
    footer_row = [
        Paragraph("<b>Total</b>", styles['TableData']),
        '',
        Paragraph(f"<b>{total_items}</b>", styles['TableNumber']),
        '',
        Paragraph(f"<b>{total_val:,.2f}</b>", styles['TableNumber']),
    ]
    data.append(footer_row)

    # Table Styling
    col_widths = [3.0*inch, 1.5*inch, 0.8*inch, 1.0*inch, 1.2*inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    
    # Premium Table Style
    t_style = TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4e73df')), # eDuka Blue
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e3e6f0')),
        
        # Alignment for numbers (columns 2,3,4) - Indices 2, 3, 4
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'), 
        
        # Footer
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#eaecf4')),
        ('TopPadding', (0, -1), (-1, -1), 10),
        ('BottomPadding', (0, -1), (-1, -1), 10),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#4e73df')),
    ])
    
    # Zebra Stripping
    for i, row in enumerate(data[1:-1], start=1):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9fc'))

    table.setStyle(t_style)
    elements.append(table)

    # --- Footer Note ---
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(f"Generated by eDuka System on {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Footer']))

    doc.build(elements)
    
    return response

def export_stock_csv(request):
    """
    Export stock data to CSV - Fallback and universally compatible format.
    """
    if not request.user.is_authenticated:
        return redirect('login')
    
    shop = None
    if getattr(request.user, 'shop', None):
        shop = request.user.shop
    elif hasattr(request.user, 'shops') and request.user.shops.exists():
        shop = request.user.shops.first()
    elif hasattr(request.user, 'employee_profile'):
        shop = request.user.employee_profile.shop
    
    if not shop:
        return redirect('stock_list')

    stocks = Stock.objects.filter(branch__shop=shop).select_related('product', 'branch')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Stock_Report_{datetime.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Category', 'Branch', 'Quantity', 'Cost Price', 'Selling Price', 'Total Value'])

    total_value_sum = 0
    for stock in stocks:
        total_val = stock.quantity * (stock.product.cost_price or 0)
        total_value_sum += total_val
        
        writer.writerow([
            stock.product.name,
            stock.product.category.name if stock.product.category else "-",
            stock.branch.name,
            stock.quantity,
            stock.product.cost_price or 0,
            stock.product.selling_price or 0,
            total_val
        ])
    
    writer.writerow([])
    writer.writerow(['', '', '', '', '', 'Total Inventory Value:', total_value_sum])

    return response


def export_stock_excel(request):
    # Security: Verify user has access to shop
    if not request.user.is_authenticated:
        return redirect('login')
    
    # Logic to get shop (duplicated from BaseShopView for function view)
    shop = None
    if getattr(request.user, 'shop', None):
        shop = request.user.shop
    elif hasattr(request.user, 'shops') and request.user.shops.exists():
        shop = request.user.shops.first()
    elif hasattr(request.user, 'employee_profile'):
        shop = request.user.employee_profile.shop
    
    if not shop:
        messages.error(request, "No shop associated.")
        return redirect('stock_list')

    stocks = Stock.objects.filter(branch__shop=shop).select_related('product', 'branch')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Stock_Report.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock List"

    # Logo
    logo_path = os.path.join(settings.BASE_DIR, 'eduka_backend', 'static', 'img', 'logoeduka.png')
    if os.path.exists(logo_path):
        try:
            img = ExcelImage(logo_path)
            img.height = 60
            img.width = 150 # Aspect ratio?
            ws.add_image(img, 'A1')
            ws.row_dimensions[1].height = 50
        except Exception as e:
            print(f"Error adding logo to Excel: {e}")

    # Headers (Start at row 5 usually to leave space for logo/title)
    ws['A4'] = f"Stock Report - {shop.name}"
    ws['A4'].font = Font(bold=True, size=14)
    ws.merge_cells('A4:F4')

    headers = ['Product Name', 'Category', 'Branch', 'Quantity', 'Cost Price', 'Selling Price', 'Total Value']
    header_row = 6
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4e73df", end_color="4e73df", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Data
    row_num = 7
    total_value_sum = 0
    for stock in stocks:
        ws.cell(row=row_num, column=1, value=stock.product.name)
        ws.cell(row=row_num, column=2, value=stock.product.category.name if stock.product.category else "-")
        ws.cell(row=row_num, column=3, value=stock.branch.name)
        ws.cell(row=row_num, column=4, value=stock.quantity)
        ws.cell(row=row_num, column=5, value=stock.product.cost_price)
        ws.cell(row=row_num, column=6, value=stock.product.selling_price)
        
        total_val = stock.quantity * (stock.product.cost_price or 0)
        ws.cell(row=row_num, column=7, value=total_val)
        total_value_sum += total_val
        
        row_num += 1

    # Footer
    ws.cell(row=row_num+1, column=6, value="Total Inventory Value:").font = Font(bold=True)
    ws.cell(row=row_num+1, column=7, value=total_value_sum).font = Font(bold=True)

    # Column/Row Resizing
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response

class StockTransferView(BaseShopView, View):
    template_name = 'inventory/stock_transfer.html'

    def get(self, request):
        shop = self.get_shop()
        if not shop:
             messages.error(request, "No shop associated.")
             return redirect('dashboard')
             
        form = StockTransferForm(shop=shop)
        
        # Get recent transfers (movements where reason implies transfer or specific type if we had one)
        # For now, just listing movements that are ADD or REDUCE might be too noisy. 
        # Ideally we should strictly identify transfers, but based on current model, we'll just show recent movements with 'Transfer' in reason? 
        # Or better separate Transfer LOGS. For now let's just show form.
        # Enhancing: Let's assume user manually writes "Transfer" in reason for now, or we define a type later if needed.
        # But wait, plan said "Log both movements".
        
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        shop = self.get_shop()
        form = StockTransferForm(request.POST, shop=shop)
        
        if form.is_valid():
            product = form.cleaned_data['product']
            source_branch = form.cleaned_data['source_branch']
            dest_branch = form.cleaned_data['destination_branch']
            qty = form.cleaned_data['quantity']
            note = form.cleaned_data['note']

            if source_branch == dest_branch:
                messages.error(request, "Source and Destination branches must be different.")
                return render(request, self.template_name, {'form': form})

            with transaction.atomic():
                # 1. Deduct from Source
                source_stock, _ = Stock.objects.get_or_create(product=product, branch=source_branch)
                
                if source_stock.quantity < qty:
                    messages.error(request, f"Insufficient stock at {source_branch.name}. Available: {source_stock.quantity}")
                    return render(request, self.template_name, {'form': form})

                source_stock.quantity -= qty
                source_stock.save()
                
                StockMovement.objects.create(
                    stock=source_stock,
                    product=product,
                    branch=source_branch,
                    quantity_change=-qty,
                    movement_type=StockMovement.Type.REDUCE, # Or custom 'TRANSFER_OUT' if enum allowed
                    reason=f"Transfer TO {dest_branch.name}: {note}",
                    user=request.user
                )

                # 2. Add to Destination
                dest_stock, _ = Stock.objects.get_or_create(product=product, branch=dest_branch)
                dest_stock.quantity += qty
                dest_stock.save()

                StockMovement.objects.create(
                     stock=dest_stock,
                    product=product,
                    branch=dest_branch,
                    quantity_change=qty,
                    movement_type=StockMovement.Type.ADD, # Or custom 'TRANSFER_IN'
                    reason=f"Transfer FROM {source_branch.name}: {note}",
                    user=request.user
                )

                messages.success(request, f"Successfully transferred {qty} {product.si_unit or 'units'} of {product.name} from {source_branch.name} to {dest_branch.name}")
                return redirect('stock_management')
        
        messages.error(request, "Please correct the errors below.")
        return render(request, self.template_name, {'form': form})

class InventoryHealthView(BaseShopView, ListView):
    template_name = 'inventory/inventory_health.html'
    context_object_name = 'low_stocks'

    def get_queryset(self):
        shop = self.get_shop()
        if shop:
            return Stock.objects.filter(branch__shop=shop, quantity__lte=F('low_stock_threshold')).select_related('product', 'branch')
        return Stock.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        shop = self.get_shop()
        if shop:
            total_items = Product.objects.filter(shop=shop, product_type=Product.Type.GOODS).count()
            stocks = Stock.objects.filter(branch__shop=shop)
            out_of_stock = stocks.filter(quantity=0).count()
            low_stock = stocks.filter(quantity__lte=F('low_stock_threshold'), quantity__gt=0).count()
            
            context['total_products'] = total_items
            context['out_of_stock_count'] = out_of_stock
            context['low_stock_count'] = low_stock
            context['healthy_stock_count'] = max(0, total_items - out_of_stock - low_stock) # Approximation
        return context

class InventoryAgingView(BaseShopView, ListView):
    template_name = 'inventory/inventory_aging.html'
    context_object_name = 'aging_stocks'

    def get_queryset(self):
        # In a real system, we'd use 'last_restock_date' or 'last_sale_date'.
        # For now, using product.updated_at as a proxy for "last activity"
        shop = self.get_shop()
        if shop:
            return Product.objects.filter(shop=shop, product_type=Product.Type.GOODS).order_by('updated_at')
        return Product.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add logic to categorize into 0-30, 31-60, 61-90, 90+ buckets logic usually handled in template or annotation
        # For simplicity here, passing the queryset which has date.
        return context

class ABCAnalysisView(BaseShopView, ListView):
    template_name = 'inventory/abc_analysis.html'
    context_object_name = 'abc_items'

    def get_queryset(self):
        shop = self.get_shop()
        if not shop: return []
        
        # ABC Logic: Value = qty * cost
        # Group by product across all branches
        products = Product.objects.filter(shop=shop, product_type=Product.Type.GOODS).annotate(
            total_qty=Sum('stocks__quantity'),
            total_value=Sum(F('stocks__quantity') * F('cost_price'))
        ).filter(total_qty__gt=0).order_by('-total_value')
        
        # Calculate Cumulative %
        total_inventory_value = sum(p.total_value or 0 for p in products)
        if total_inventory_value == 0: return []
        
        running_sum = 0
        abc_list = []
        for p in products:
            val = p.total_value or 0
            running_sum += val
            cum_percent = (running_sum / total_inventory_value) * 100
            
            if cum_percent <= 70:
                grade = 'A'
            elif cum_percent <= 90:
                grade = 'B'
            else:
                grade = 'C'
            
            p.abc_grade = grade
            p.share_percent = (val / total_inventory_value) * 100
            abc_list.append(p)
            
        return abc_list

from django.db.models.functions import Coalesce

class ProfitabilityReportView(BaseShopView, ListView):
    template_name = 'inventory/profitability_report.html'
    context_object_name = 'profit_items'

    def get_queryset(self):
        shop = self.get_shop()
        if shop:
             # Advanced Profitability:
             # 1. Total Qty Sold
             # 2. Total Revenue (Qty * Sold Price)
             # 3. Estimated COGS (Qty * Current Cost Price) - Note: FIFO/LIFO is complex, using replacement cost for now.
             
             return Product.objects.filter(shop=shop, product_type=Product.Type.GOODS).annotate(
                 total_qty_sold=Coalesce(Sum('saleitem__quantity'), 0),
                 total_revenue=Coalesce(Sum(F('saleitem__quantity') * F('saleitem__price')), Decimal('0.00')),
             ).annotate(
                 estimated_cogs=F('total_qty_sold') * F('cost_price'),
                 gross_profit=F('total_revenue') - (F('total_qty_sold') * F('cost_price')),
                 margin=F('selling_price') - F('cost_price') # Theoretical Unit Margin
             ).filter(total_qty_sold__gt=0).order_by('-gross_profit')
        return Product.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Prepare Chart Data
        products = list(context['profit_items'][:10]) # Top 10
        context['chart_labels'] = [p.name for p in products]
        context['chart_data'] = [float(p.gross_profit) for p in products]
        return context

class ProductImportView(BaseShopView, TemplateView):
    template_name = 'inventory/product_import.html'
    
    def post(self, request, *args, **kwargs):
        file = request.FILES.get('file')
        if not file:
            messages.error(request, "No file uploaded.")
            return redirect('product_import')

        if not file.name.endswith('.csv'):
            messages.error(request, "Please upload a valid CSV file.")
            return redirect('product_import')

        shop = self.get_shop()
        if not shop:
             messages.error(request, "No shop found. Please create a shop first.")
             return redirect('dashboard')

        try:
            # Read and decode the file
            decoded_file = file.read().decode('utf-8-sig') # utf-8-sig handles BOM if present (Excel often adds it)
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            # Basic validation of headers (checking only a key one to ensure compatibility)
            # Headers from template: 'Name (Jina)', 'Category (Kundi)', etc.
            field_names = reader.fieldnames
            if not field_names or 'Name (Jina)' not in field_names:
                 messages.error(request, "Invalid CSV format. Please use the provided template.")
                 return redirect('product_import')

            created_count = 0
            updated_count = 0
            errors = []

            with transaction.atomic():
                for index, row in enumerate(reader, start=1):
                    try:
                        name = row.get('Name (Jina)', '').strip()
                        if not name: continue # Skip empty rows
                        
                        category_name = row.get('Category (Kundi)', '').strip()
                        product_type_str = row.get('Type (Bidhaa/Huduma - GOODS/SERVICE)', 'GOODS').upper()
                        sku = row.get('SKU', '').strip()
                        barcode = row.get('Barcode', '').strip()
                        
                        # Price parsing
                        try:
                            selling_price = Decimal(row.get('Selling Price (Bei Kuuzia)', '0').replace(',', '') or 0)
                            cost_price = Decimal(row.get('Cost Price (Bei Kununua)', '0').replace(',', '') or 0)
                        except:
                            errors.append(f"Row {index}: Invalid price format.")
                            continue

                        si_unit = row.get('SI Unit (Kipimo)', '').strip()
                        opening_stock = int(row.get('Opening Stock (Stock)', '0') or 0)
                        threshold = int(row.get('Low Stock Threshold (Kiwango cha chini)', '0') or 0)

                        # Logic:
                        # 1. Get/Create Category
                        category = None
                        if category_name:
                            category, _ = Category.objects.get_or_create(shop=shop, name=category_name)
                        
                        # 2. Update or Create Product (Match by SKU first, then Name if SKU empty)
                        product = None
                        if sku:
                            product = Product.objects.filter(shop=shop, sku=sku).first()
                        
                        if not product:
                            product = Product.objects.filter(shop=shop, name=name).first()

                        product_type = Product.Type.SERVICE if 'SERVICE' in product_type_str else Product.Type.GOODS

                        if product:
                            # Update existing
                            product.category = category
                            product.selling_price = selling_price
                            product.cost_price = cost_price
                            product.si_unit = si_unit
                            product.product_type = product_type
                            if barcode: product.barcode = barcode
                            product.save()
                            updated_count += 1
                        else:
                            # Create new
                            product = Product.objects.create(
                                shop=shop,
                                name=name,
                                category=category,
                                product_type=product_type,
                                sku=sku,
                                barcode=barcode,
                                selling_price=selling_price,
                                cost_price=cost_price,
                                si_unit=si_unit
                            )
                            created_count += 1

                        # 3. Handle Stock (Only for Goods)
                        if product.product_type == Product.Type.GOODS:
                            # Update stock for the main branch (or all branches?)
                            # For simplicity in import, apply to the first found branch or all.
                            # Usually imports are for "Opening Stock" which implies setting the base level.
                            branches = Branch.objects.filter(shop=shop)
                            for branch in branches:
                                stock, created = Stock.objects.get_or_create(product=product, branch=branch)
                                # If import specifies stock, do we ADD or SET?
                                # Usually import = SET current state.
                                # But if multiple branches exist, this is ambiguous.
                                # Let's assume this import sets the quantity for the MAIN branch if it exists, or random one.
                                if branch.is_main or branches.count() == 1:
                                    stock.quantity += opening_stock
                                    stock.low_stock_threshold = threshold
                                    stock.save()
                                    
                                    # Log Movement for Import
                                    if opening_stock > 0:
                                        StockMovement.objects.create(
                                            stock=stock,
                                            product=product,
                                            branch=branch,
                                            quantity_change=opening_stock,
                                            movement_type=StockMovement.Type.ADD,
                                            reason="Bulk Import: Added via CSV",
                                            user=request.user
                                        )

                    except Exception as e:
                        errors.append(f"Row {index} ({name}): {str(e)}")

            if errors:
                messages.warning(request, f"Import completed with warnings. Created: {created_count}, Updated: {updated_count}. Errors: {len(errors)}")
                # Optionally show first few errors
                for err in errors[:5]:
                    messages.error(request, err)
            else:
                messages.success(request, f"Successfully imported {created_count} new products and updated {updated_count} existing products.")

        except Exception as e:
            messages.error(request, f"Critical error during import: {str(e)}")
        
        return redirect('product_list')

class PurchaseCreateView(BaseShopView, View):
    template_name = 'inventory/purchase_form.html'

    def get(self, request):
        shop = self.get_shop()
        if not shop:
             messages.error(request, "No shop associated.")
             return redirect('dashboard')
        
        form = PurchaseForm(shop=shop)
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        shop = self.get_shop()
        form = PurchaseForm(request.POST, shop=shop)
        
        if form.is_valid():
            product = form.cleaned_data['product']
            branch = form.cleaned_data['branch']
            qty = form.cleaned_data['quantity']
            cost_price = form.cleaned_data['cost_price']
            supplier = form.cleaned_data['supplier']
            ref_no = form.cleaned_data['reference_number']
            note = form.cleaned_data['note']

            try:
                with transaction.atomic():
                    # 1. Update/Create Stock
                    stock, _ = Stock.objects.get_or_create(product=product, branch=branch)
                    stock.quantity += qty
                    stock.save()

                    # 2. Update Product Cost Price (Weighted Average or Last Price?)
                    # Requirement says "Update cost price if provided".
                    # Let's assume replacement (Last Price) for now as it's common in simple systems,
                    # or just update the field.
                    if cost_price:
                        product.cost_price = cost_price
                        product.save()

                    # 3. Log Movement
                    reason_text = f"Purchase from {supplier}" if supplier else "Purchase"
                    if ref_no: reason_text += f" (Ref: {ref_no})"
                    if note: reason_text += f" - {note}"

                    StockMovement.objects.create(
                        stock=stock,
                        product=product,
                        branch=branch,
                        quantity_change=qty,
                        movement_type=StockMovement.Type.PURCHASE,
                        reason=reason_text,
                        user=request.user
                    )

                    messages.success(request, f"Purchase recorded: {qty} x {product.name}")
                    return redirect('purchase_list')

            except Exception as e:
                messages.error(request, f"Error recording purchase: {str(e)}")
        
        return render(request, self.template_name, {'form': form})

class PurchaseListView(BaseShopView, ListView):
    model = StockMovement
    template_name = 'inventory/purchase_list.html'
    context_object_name = 'purchases'
    paginate_by = 20

    def get_queryset(self):
        shop = self.get_shop()
        if shop:
            return StockMovement.objects.filter(
                branch__shop=shop, 
                movement_type=StockMovement.Type.PURCHASE
            ).select_related('product', 'branch', 'user').order_by('-created_at')
        return StockMovement.objects.none()

class PurchaseRecentView(BaseShopView, ListView):
    model = StockMovement
    template_name = 'inventory/purchase_recent.html'
    context_object_name = 'recent_purchases'
    
    def get_queryset(self):
        shop = self.get_shop()
        if shop:
            return StockMovement.objects.filter(
                branch__shop=shop, 
                movement_type=StockMovement.Type.PURCHASE
            ).select_related('product', 'branch', 'user').order_by('-created_at')[:10]
        return StockMovement.objects.none()

class ProductTemplateDownloadView(LoginRequiredMixin, View):
    def get(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="product_import_template.csv"'

        writer = csv.writer(response)
        # Headers matching the requested system format
        headers = [
            'Name (Jina)', 
            'Category (Kundi)', 
            'Type (Bidhaa/Huduma - GOODS/SERVICE)', 
            'SKU', 
            'Barcode', 
            'Selling Price (Bei Kuuzia)', 
            'Cost Price (Bei Kununua)', 
            'SI Unit (Kipimo)', 
            'Opening Stock (Stock)', 
            'Low Stock Threshold (Kiwango cha chini)'
        ]
        writer.writerow(headers)
        
        # Add an example row
        writer.writerow([
            'Coca Cola 500ml', 
            'Drinks', 
            'GOODS', 
            'CC500', 
            '123456789', 
            '1000', 
            '800', 
            'Pcs', 
            '50', 
            '10'
        ])
        
        return response

class BarcodePrintView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Get product IDs from query param `ids` (comma separated)
        product_ids_str = request.GET.get('ids', '')
        if not product_ids_str:
             # Fallback: check if single product_id is passed in kwargs (e.g. /barcode/<id>/)
            single_id = kwargs.get('pk')
            if single_id:
                product_ids = [single_id]
            else:
                return HttpResponse("No products selected", status=400)
        else:
            product_ids = [int(id) for id in product_ids_str.split(',') if id.isdigit()]
            
        products = Product.objects.filter(id__in=product_ids)
        if not products.exists():
            return HttpResponse("Products not found", status=404)
            
        pdf_buffer = generate_pdf_labels(products)
        
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="barcodes.pdf"'
        return response
